import base64, io, json
from dash.dependencies import Output, Input, State
from dash import html, dcc, dash_table
import dash_bootstrap_components as dbc
import pandas as pd
from app import *

import dash_uploader as du
import uuid
import os
import csv
import xlrd, xlwt
import openpyxl
import json


# 将csv文件转换为json数据类型
def convert_import_csv(filePath):
    with open(filePath, 'r') as file:
        reader = csv.reader(file)
        data = [row for row in reader]
    return data


# 将xls文件转换为json数据类型
def convert_import_xls(filePath):
    workbook = xlrd.open_workbook(filePath)
    sheet = workbook.sheet_by_index(0)
    data = []
    for row in range(sheet.nrows):
        data.append(sheet.row_values(row))
    return data


# 将xlsx文件转换为json数据类型
def convert_import_xlsx(filePath):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append([i if i != "NA" else None for i in row])
    return data


# 将json文件转换为json数据类型
def convert_import_json(filePath):
    with open(filePath, 'r') as file:
        data = json.load(file)
    return data


# 将json数据类型转换为csv文件
def convert_export_csv(filePath, data):
    with open(filePath, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)


# 将json数据类型转换为xls文件
def convert_export_xls(filePath, data):
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')
    for row_index, row_data in enumerate(data):
        for col_index, cell_data in enumerate(row_data):
            sheet.write(row_index, col_index, cell_data)
    workbook.save(filePath)


# 将json数据类型转换为xlsx文件
def convert_export_xlsx(filePath, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_data in data:
        sheet.append(row_data)
    workbook.save(filePath)


# 将json数据类型转换为json文件
def convert_export_json(filePath, data):
    with open(filePath, 'w+') as file:
        file.write(json.dumps(data, ensure_ascii=False))

from flask import Flask, request

@server.route('/DataFormatConvert/convert', methods=['POST'])
def DataFormatConvert_convert():
    if request.content_type.startswith('multipart/form-data'):
        fileuuid = request.form.get('uuid')
        filename = request.form.get('filename')
        convertto = request.form.get('convertto')
    else:
        fileuuid = request.values.get('uuid')
        filename = request.values.get("filename")
        convertto = request.values.get('convertto')
         
    if './' in fileuuid or './' in filename:
        return
    
    filepath = f"{PATH}/upload/{fileuuid}/{filename}"
    
    if filename.endswith("csv"):
        data = convert_import_csv(filepath)
    elif filename.endswith("xls"):
        data = convert_import_xls(filepath)
    elif filename.endswith("xlsx"):
        data = convert_import_xlsx(filepath)
    elif filename.endswith("json"):
        data = convert_import_json(filepath)
    else:
        return "not valid format"
    
    if convertto == "csv":
        converttoPath = f"{PATH}/upload/{fileuuid}/{os.path.splitext(filename)[0]}_convert.csv"
        convert_export_csv(converttoPath, data)
    elif convertto == "xls":
        converttoPath = f"{PATH}/upload/{fileuuid}/{os.path.splitext(filename)[0]}_convert.xls"
        convert_export_xls(converttoPath, data)
    elif convertto == "xlsx":
        converttoPath = f"{PATH}/upload/{fileuuid}/{os.path.splitext(filename)[0]}_convert.xlsx"
        convert_export_xlsx(converttoPath, data)
    elif convertto == "json":
        converttoPath = f"{PATH}/upload/{fileuuid}/{os.path.splitext(filename)[0]}_convert.json"
        convert_export_json(converttoPath, data)
    else:
        return "convertto error"
        
    directory, filename = os.path.split(converttoPath)
    response = make_response(send_from_directory(directory, filename, as_attachment=True))
    response.headers["Content-Disposition"] = "attachment; filename={}".format(filename.encode().decode('latin-1'))
    return response
    


du.configure_upload(app, folder='/www/wwwroot/DataProcessing/upload')

    
def Layout():
    return dbc.Container([
        dbc.Container([
            html.Br(),
            html.H1("PDFC - Python-based Data Format Convert"),
            html.Br(),
            du.Upload(
                id='[DataFormatConvert/DataFormatConvert]uploader',
                text='Click or drag the file here to upload',
                text_completed='File: ',
                cancel_button=True,
                pause_button=True,
                filetypes=['csv', 'xls', 'xlsx', 'json'],
                default_style={
                    'background-color': '#fafafa',
                    'font-weight': 'bold'
                },
                upload_id=uuid.uuid1()
            ),
            html.Br(),
            dbc.Col([
                html.H3("Convert to"),
                dbc.Select(
                    id="[DataFormatConvert/DataFormatConvert]ToCheckList",
                    options=["csv", "xls", "xlsx", "json"],
                    value="csv"
                ),
            ]),
            html.Br(),
            
            html.Form([
                dcc.Input(id='[DataFormatConvert/DataFormatConvert]uuid', name='uuid', style={"display": "none"}),
                dcc.Input(id='[DataFormatConvert/DataFormatConvert]filename', name='filename', style={"display": "none"}),
                dcc.Input(id='[DataFormatConvert/DataFormatConvert]convertto', name='convertto', style={"display": "none"}),
                html.Br(),
                html.Div([dbc.Button("Convert")], style={"text-align": "right"}),
            ], method="POST", action="/DataFormatConvert/convert"),
        ]),
    
    ], fluid=True)


@app.callback(
    Output('[DataFormatConvert/DataFormatConvert]uuid', 'value'),
    Output('[DataFormatConvert/DataFormatConvert]filename', 'value'),
    Input('[DataFormatConvert/DataFormatConvert]uploader', 'isCompleted'),
    State('[DataFormatConvert/DataFormatConvert]uploader', 'fileNames'),
    State('[DataFormatConvert/DataFormatConvert]uploader', 'upload_id'),
)
def show_upload_status(isCompleted, fileNames, upload_id):
    if isCompleted:
        return upload_id, fileNames[0]

    return dash.no_update, dash.no_update


@app.callback(
    Output('[DataFormatConvert/DataFormatConvert]convertto', 'value'),
    Input('[DataFormatConvert/DataFormatConvert]ToCheckList', 'value')
)
def show_upload_status(convertto):
    return convertto

